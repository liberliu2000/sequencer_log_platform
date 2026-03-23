from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from app.core.settings import get_settings
from app.services.query_service import QueryService


class ExportService:
    def __init__(self, db: Session):
        self.db = db
        self.settings = get_settings()
        self.query = QueryService(db)

    def export_events_csv(self, task_id: int, task_uuid: str) -> str:
        rows = self.query.list_events(task_id=task_id, limit=100000)
        output = Path(self.settings.export_dir) / f"{task_uuid}_events.csv"
        self._write_csv(output, rows)
        return str(output)

    def export_error_report_csv(self, task_id: int, task_uuid: str) -> str:
        rows = self.query.get_error_clusters(task_id=task_id)
        output = Path(self.settings.export_dir) / f"{task_uuid}_errors.csv"
        self._write_csv(output, rows)
        return str(output)

    def export_json_report(self, task_id: int, task_uuid: str) -> str:
        payload = self._build_report_payload(task_id)
        output = Path(self.settings.export_dir) / f"{task_uuid}_report.json"
        output.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        return str(output)

    def export_excel_report(self, task_id: int, task_uuid: str) -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError as exc:
            raise RuntimeError("导出 Excel 依赖缺失：请安装 openpyxl（python -m pip install openpyxl）") from exc

        payload = self._build_report_payload(task_id)
        output = Path(self.settings.export_dir) / f"{task_uuid}_report.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard"
        dash = payload["dashboard"]
        ws.append(["指标", "值"])
        for k, v in dash.items():
            if isinstance(v, list):
                continue
            ws.append([k, self._safe_value(v)])
        self._autosize_sheet(ws, get_column_letter)
        self._add_sheet(wb, "Errors", payload["errors"], get_column_letter)
        self._add_sheet(wb, "Cycles", payload["cycle_summary"], get_column_letter)
        self._add_sheet(wb, "Audit", payload["audit_logs"], get_column_letter)
        self._add_sheet(wb, "LLM", payload["llm_results"], get_column_letter)
        wb.save(output)
        return str(output)

    def export_pdf_report(self, task_id: int, task_uuid: str) -> str:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.utils import simpleSplit
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            from reportlab.pdfgen import canvas
        except ModuleNotFoundError as exc:
            raise RuntimeError("导出 PDF 依赖缺失：请安装 reportlab（python -m pip install reportlab）") from exc

        payload = self._build_report_payload(task_id)
        output = Path(self.settings.export_dir) / f"{task_uuid}_report.pdf"
        c = canvas.Canvas(str(output), pagesize=A4)
        width, height = A4
        margin_x = 40
        y = height - 40
        title_font, body_font = self._register_pdf_fonts(pdfmetrics, UnicodeCIDFont)
        c.setTitle(f"Sequencer Log Report - {task_uuid}")
        y = self._draw_wrapped_line(c, f"测序仪日志分析报告 / Sequencer Log Report - {task_uuid}", margin_x, y, width - 2 * margin_x, simpleSplit, title_font, 14, 18)
        dash = payload["dashboard"]
        lines = [
            "",
            "一、项目概览",
            f"文件数: {dash.get('file_count', 0)}",
            f"总事件数: {dash.get('total_events', 0)}",
            f"总错误数: {dash.get('total_errors', 0)}",
            f"唯一错误数: {dash.get('unique_error_count', 0)}",
            "",
            "二、Top 错误簇",
        ]
        for row in payload["errors"][:10]:
            lines.append(f"- {row.get('display_signature') or row.get('representative_message') or ''} (count={row.get('count', 0)}, family={row.get('error_family') or '未知'})")
        lines += ["", "三、Cycle 总耗时"]
        for row in payload["cycle_summary"][:10]:
            lines.append(f"- Cycle {row.get('cycle_no')} | chip={row.get('chip_name') or '-'} | duration={row.get('total_duration_ms')} ms | start={row.get('started_at_text') or row.get('started_at') or '-'} | end={row.get('ended_at_text') or row.get('ended_at') or '-'}")
        lines += ["", "四、最近审计日志"]
        for row in payload["audit_logs"][:10]:
            lines.append(f"- {row.get('created_at', '')} | {row.get('action', '')} | {row.get('status', '')} | {row.get('stage') or '-'} | {row.get('detail') or ''}")
        c.setFont(body_font, 10)
        for line in lines:
            y = self._draw_wrapped_line(c, line, margin_x, y, width - 2 * margin_x, simpleSplit, body_font, 10, 14)
            if y < 60:
                c.showPage()
                c.setFont(body_font, 10)
                y = height - 40
        c.save()
        return str(output)

    def _build_report_payload(self, task_id: int) -> dict:
        return {
            "dashboard": self.query.get_dashboard(task_id),
            "errors": self.query.get_error_clusters(task_id),
            "cycle_summary": self.query.get_cycle_summaries(task_id),
            "audit_logs": self.query.get_audit_logs(task_id),
            "llm_results": self.query.get_llm_results(task_id),
        }

    @staticmethod
    def _safe_value(value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, (str, int, float, bool)):
            return value
        if isinstance(value, (list, tuple, dict)):
            try:
                return json.dumps(value, ensure_ascii=False, default=str)
            except Exception:
                return str(value)
        return str(value)

    @classmethod
    def _write_csv(cls, output: Path, rows: list[dict]) -> None:
        if not rows:
            output.write_text("", encoding="utf-8")
            return
        with output.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            for row in rows:
                writer.writerow({k: cls._safe_value(v) for k, v in row.items()})

    @classmethod
    def _add_sheet(cls, wb, name: str, rows: list[dict], get_column_letter) -> None:
        ws = wb.create_sheet(title=name[:31])
        if not rows:
            ws.append(["empty"])
            return
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([cls._safe_value(row.get(h)) for h in headers])
        cls._autosize_sheet(ws, get_column_letter)

    @staticmethod
    def _autosize_sheet(ws, get_column_letter) -> None:
        max_widths: dict[int, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                value = "" if cell.value is None else str(cell.value)
                max_widths[cell.column] = max(max_widths.get(cell.column, 0), min(len(value), 60))
        for idx, width in max_widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = max(12, min(width + 2, 64))

    @staticmethod
    def _register_pdf_fonts(pdfmetrics, UnicodeCIDFont) -> tuple[str, str]:
        try:
            pdfmetrics.getFont("STSong-Light")
        except KeyError:
            pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        return "STSong-Light", "STSong-Light"

    @staticmethod
    def _draw_wrapped_line(c, text_line: str, x: float, y: float, max_width: float, simpleSplit, font_name: str, font_size: int, leading: int = 14) -> float:
        text_line = text_line or ""
        wrapped = simpleSplit(str(text_line), font_name, font_size, max_width) or [""]
        c.setFont(font_name, font_size)
        for seg in wrapped:
            c.drawString(x, y, seg)
            y -= leading
        return y
